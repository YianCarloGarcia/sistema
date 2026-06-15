from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.mail import send_mail
from django.db.models import Q
from django.conf import settings
import io, csv, zipfile, json, os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Estudiante, Asistencia
from .forms import EstudianteForm, UsuarioCrearForm, UsuarioEditarForm


# ── Decoradores de rol ──────────────────────────────────────

def solo_directivo(view_func):
    """Bloquea el acceso a usuarios que no sean directivos ni superusuarios."""
    def wrapper(request, *args, **kwargs):
        if not _es_directivo(request.user):
            messages.error(request, "No tiene permisos para acceder a esta sección.")
            return redirect('inicio')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def _es_directivo(user):
    return user.is_superuser or user.groups.filter(name='directivo').exists()


def _es_docente(user):
    return user.groups.filter(name='docente').exists()


def _asegurar_grupos():
    """Crea los grupos Docente y Directivo si no existen."""
    Group.objects.get_or_create(name='docente')
    Group.objects.get_or_create(name='directivo')


# ── Email helpers ───────────────────────────────────────────

def _enviar_bienvenida(user, password_plano, rol):
    rol_label = 'Directivo' if rol == 'directivo' else 'Docente'
    permisos  = ('Acceso completo al sistema.' if rol == 'directivo'
                 else 'Puede ver y buscar estudiantes. No puede editar ni eliminar.')
    try:
        send_mail(
            subject='Bienvenido al Sistema Escolar — Sus credenciales',
            message=(
                f"Hola {user.get_full_name() or user.username},\n\n"
                f"Se ha creado su cuenta en el Sistema de Gestión Escolar.\n\n"
                f"Usuario:     {user.username}\n"
                f"Contraseña:  {password_plano}\n"
                f"Rol:         {rol_label}\n"
                f"Permisos:    {permisos}\n\n"
                f"Por seguridad, cambie su contraseña después de iniciar sesión.\n"
                f"Para restablecer su contraseña, use la opción '¿Olvidó su contraseña?' en la página de inicio de sesión.\n\n"
                f"Sistema Escolar"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
        return True
    except Exception as e:
        return str(e)


# ── Helpers comunes ──────────────────────────────────────────

CAMPOS_EXPORTAR = [
    ('documento',           'Documento'),
    ('tipo',                'Tipo Doc.'),
    ('apellidos',           'Apellidos'),
    ('nombres',             'Nombres'),
    ('jornada',             'Jornada'),
    ('curso',               'Curso'),
    ('linea',               'Linea'),
    ('celular',             'Celular'),
    ('email',               'Email'),
    ('acudiente',           'Acudiente'),
    ('parentesco',          'Parentesco'),
    ('tel_acudiente',       'Tel. Acudiente'),
    ('tel2_acudiente',      'Tel. Acudiente 2'),
    ('direccion',           'Direccion'),
    ('ocupacion_acudiente', 'Ocupacion Acudiente'),
    ('eps',                 'EPS'),
    ('observaciones',       'Observaciones'),
    ('archivo_foto',        'Archivo Foto'),
]
CAMPOS_IMPORTAR_REQUERIDOS = ['documento', 'apellidos', 'nombres', 'jornada', 'curso', 'linea']
VALORES_JORNADA = [v for v, _ in Estudiante.JORNADA]
VALORES_LINEA   = [v for v, _ in Estudiante.LINEA_MEDIA]
VALORES_TIPO    = [v for v, _ in Estudiante.TIPOS_DOCUMENTO]


def _qs_filtrada(get):
    qs = Estudiante.objects.all().order_by('apellidos', 'nombres')
    q       = get.get('q', '').strip()
    jornadas = get.getlist('jornada')
    grados   = get.getlist('grado')
    lineas   = get.getlist('linea')
    cursos   = get.getlist('curso')
    if q:
        qs = qs.filter(Q(nombres__icontains=q) | Q(apellidos__icontains=q) | Q(documento__icontains=q))
    if jornadas:
        qs = qs.filter(jornada__in=jornadas)
    if grados:
        q_grado = Q()
        for g in grados:
            q_grado |= Q(curso__startswith=g)
        qs = qs.filter(q_grado)
    if lineas:
        qs = qs.filter(linea__in=lineas)
    if cursos:
        qs = qs.filter(curso__in=cursos)
    return qs


def _build_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    hf = PatternFill("solid", fgColor="1a3a6e")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (_, label) in enumerate(CAMPOS_EXPORTAR, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = hf; c.font = hfont; c.alignment = halign; c.border = border
    ws.row_dimensions[1].height = 28
    alt = PatternFill("solid", fgColor="F8FAFC")
    for ri, est in enumerate(queryset, 2):
        for ci, (field, _) in enumerate(CAMPOS_EXPORTAR, 1):
            if field == 'archivo_foto':
                val = os.path.basename(est.foto.name) if est.foto else ''
            else:
                val = getattr(est, field, '') or ''
            c = ws.cell(row=ri, column=ci, value=str(val))
            c.alignment = Alignment(vertical="center"); c.border = border
            if ri % 2 == 0: c.fill = alt
    anchos = [14,8,20,20,10,8,10,14,24,22,12,14,14,24,20,14,30,20]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _leer_excel(archivo):
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as e:
        return None, [f"No se pudo leer el archivo: {e}"]
    ws = wb.active
    headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(1, ws.max_column + 1)]
    label_to_field = {label: field for field, label in CAMPOS_EXPORTAR}
    col_map = {i: label_to_field[h] for i, h in enumerate(headers) if h in label_to_field}
    if 'documento' not in col_map.values():
        return None, ["El archivo no tiene columna 'Documento'. Use la plantilla descargada."]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row): continue
        fila = {field: (str(row[i]).strip() if i < len(row) and row[i] is not None else '')
                for i, field in col_map.items()}
        filas.append(fila)
    return filas, []


# ── Vistas generales ────────────────────────────────────────

@login_required
def inicio(request):
    return render(request, 'paginas/inicio.html', {
        'es_directivo': _es_directivo(request.user),
    })

@login_required
def nosotros(request):
    return render(request, 'paginas/nosotros.html')


# ── Estudiantes ──────────────────────────────────────────────

@login_required
def estudiantes(request):
    qs = _qs_filtrada(request.GET)
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    cursos_disponibles = Estudiante.objects.values_list('curso', flat=True).distinct().order_by('curso')
    return render(request, 'estudiantes/index.html', {
        'page_obj':           page_obj,
        'total_filtrado':     qs.count(),
        'total_general':      Estudiante.objects.count(),
        'cursos_disponibles': cursos_disponibles,
        'jornadas':           Estudiante.JORNADA,
        'lineas':             Estudiante.LINEA_MEDIA,
        'filtro_q':        request.GET.get('q', ''),
        'filtro_jornadas': request.GET.getlist('jornada'),
        'filtro_grados':   request.GET.getlist('grado'),
        'filtro_lineas':   request.GET.getlist('linea'),
        'hay_filtros':     bool(request.GET.get('q') or request.GET.getlist('jornada') or
                               request.GET.getlist('grado') or request.GET.getlist('linea')),
        'es_directivo':   _es_directivo(request.user),
    })


@login_required
@solo_directivo
def crear(request):
    formulario = EstudianteForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        messages.success(request, "Estudiante creado correctamente.")
        return redirect('estudiantes')
    return render(request, 'estudiantes/crear.html', {'formulario': formulario})


@login_required
@solo_directivo
def editar(request, id):
    estudiante = get_object_or_404(Estudiante, id=id)
    formulario = EstudianteForm(request.POST or None, request.FILES or None, instance=estudiante)
    if formulario.is_valid():
        formulario.save()
        messages.success(request, "Estudiante actualizado correctamente.")
        return redirect('estudiantes')
    return render(request, 'estudiantes/editar.html', {'formulario': formulario})


@login_required
@solo_directivo
def eliminar(request, id):
    get_object_or_404(Estudiante, id=id).delete()
    messages.success(request, "Estudiante eliminado.")
    return redirect('estudiantes')


@login_required
def detalle(request, id):
    estudiante = get_object_or_404(Estudiante, id=id)
    anio = timezone.now().year
    conteos = {
        tipo: Asistencia.objects.filter(estudiante=estudiante, tipo=tipo, fecha__year=anio).count()
        for tipo in ('ALM', 'TAR', 'UNI', 'ASI')
    }
    return render(request, 'estudiantes/detalle.html', {
        'estudiante':   estudiante,
        'es_directivo': _es_directivo(request.user),
        'conteos':      conteos,
        'anio':         anio,
    })


# ── Almuerzo ─────────────────────────────────────────────────

BLOQUEO_SEGUNDOS = 5

TIPOS_ESCANER = {
    'ALM': {'label': 'Registro de Almuerzo',  'icono': '\U0001f37d\ufe0f', 'color': '#e63946'},
    'TAR': {'label': 'Llegada Tarde',          'icono': '\u23f0',            'color': '#f4a261'},
    'UNI': {'label': 'Porte de Uniforme',      'icono': '\U0001f454',        'color': '#2a9d8f'},
    'ASI': {'label': 'Asistencia a Clase',     'icono': '\U0001f4cb',        'color': '#457b9d'},
}

@login_required
def escaner(request):
    tipo_actual = request.GET.get('tipo', 'ALM')
    if tipo_actual not in TIPOS_ESCANER:
        tipo_actual = 'ALM'

    mensaje = contador = nombre = None

    if request.method == 'POST':
        documento   = request.POST.get('documento', '').strip()
        tipo_actual = request.POST.get('tipo', tipo_actual)
        if tipo_actual not in TIPOS_ESCANER:
            tipo_actual = 'ALM'

        try:
            est    = Estudiante.objects.get(documento=documento)
            ahora  = timezone.now()
            limite = ahora - timedelta(seconds=BLOQUEO_SEGUNDOS)
            ultimo = Asistencia.objects.filter(estudiante=est, tipo=tipo_actual).order_by('-fecha', '-hora').first()
            if ultimo:
                fhu = timezone.make_aware(timezone.datetime.combine(ultimo.fecha, ultimo.hora))
                if fhu > limite:
                    mensaje = f"Espere {BLOQUEO_SEGUNDOS} segundos antes de volver a escanear"
                else:
                    Asistencia.objects.create(estudiante=est, tipo=tipo_actual)
            else:
                Asistencia.objects.create(estudiante=est, tipo=tipo_actual)
            hoy      = timezone.localdate()
            contador = Asistencia.objects.filter(estudiante=est, fecha=hoy, tipo=tipo_actual).count()
            nombre   = f"{est.nombres} {est.apellidos}"
        except Estudiante.DoesNotExist:
            mensaje = "Documento no encontrado"

    info = TIPOS_ESCANER[tipo_actual]
    return render(request, 'estudiantes/escaner.html', {
        'mensaje':     mensaje,
        'contador':    contador,
        'nombre':      nombre,
        'tipo_actual': tipo_actual,
        'tipos':       TIPOS_ESCANER,
        'info':        info,
    })


# Redirige la URL antigua para no romper bookmarks
@login_required
def almuerzo(request):
    from django.shortcuts import redirect as _redirect
    return _redirect('/estudiantes/escaner/?tipo=ALM')


# ── Escáner AJAX — responde JSON, sin recargar la página ─────────────────────
@login_required
def escaner_registrar(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'mensaje': 'Método no permitido'}, status=405)

    documento   = request.POST.get('documento', '').strip()
    tipo_actual = request.POST.get('tipo', 'ALM')
    if tipo_actual not in TIPOS_ESCANER:
        tipo_actual = 'ALM'

    if not documento:
        return JsonResponse({'ok': False, 'mensaje': 'Documento vacío'})

    try:
        est    = Estudiante.objects.get(documento=documento)
        ahora  = timezone.now()
        limite = ahora - timedelta(seconds=BLOQUEO_SEGUNDOS)
        ultimo = Asistencia.objects.filter(
            estudiante=est, tipo=tipo_actual
        ).order_by('-fecha', '-hora').first()

        if ultimo:
            fhu = timezone.make_aware(
                timezone.datetime.combine(ultimo.fecha, ultimo.hora)
            )
            if fhu > limite:
                return JsonResponse({
                    'ok': False,
                    'tipo': 'espera',
                    'mensaje': f'Espere {BLOQUEO_SEGUNDOS} segundos antes de volver a escanear',
                    'nombre': f'{est.nombres} {est.apellidos}',
                })

        Asistencia.objects.create(estudiante=est, tipo=tipo_actual)
        hoy      = timezone.localdate()
        contador = Asistencia.objects.filter(
            estudiante=est, fecha=hoy, tipo=tipo_actual
        ).count()

        return JsonResponse({
            'ok':       True,
            'nombre':   f'{est.nombres} {est.apellidos}',
            'contador': contador,
            'label':    TIPOS_ESCANER[tipo_actual]['label'],
        })

    except Estudiante.DoesNotExist:
        return JsonResponse({'ok': False, 'mensaje': 'Documento no encontrado'})


def exit(request):
    logout(request)
    return redirect('inicio')


# ── Gestión masiva (solo directivos) ─────────────────────────

@login_required
@solo_directivo
def gestion_masiva(request):
    cursos_disponibles = Estudiante.objects.values_list('curso', flat=True).distinct().order_by('curso')
    return render(request, 'estudiantes/gestion_masiva.html', {
        'total':              Estudiante.objects.count(),
        'campos':             CAMPOS_EXPORTAR,
        'jornadas':           Estudiante.JORNADA,
        'lineas':             Estudiante.LINEA_MEDIA,
        'cursos_disponibles': cursos_disponibles,
    })


@login_required
@solo_directivo
def exportar_estudiantes(request):
    qs  = _qs_filtrada(request.GET)
    buf = _build_excel(qs)
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estudiantes.xlsx'
    return response


@login_required
@solo_directivo
def importar_estudiantes(request):
    if request.method == 'POST':
        if request.POST.get('confirm') == '1':
            filas = json.loads(request.POST.get('filas_json', '[]'))
            creados = actualizados = omitidos = 0
            for fila in filas:
                if fila.get('error'): omitidos += 1; continue
                doc = fila.get('documento', '').strip()
                if not doc: omitidos += 1; continue
                campos = {k: v for k, v in fila.items()
                          if k not in ('error', 'errores', 'advertencias', 'fila', 'existe', 'documento', 'archivo_foto')}
                archivo_foto = fila.get('archivo_foto', '').strip()
                if archivo_foto:
                    ruta_foto = os.path.join(settings.MEDIA_ROOT, 'fotos', archivo_foto)
                    if os.path.isfile(ruta_foto):
                        campos['foto'] = f'fotos/{archivo_foto}'
                _, creado = Estudiante.objects.update_or_create(documento=doc, defaults=campos)
                creados += 1 if creado else 0
                actualizados += 0 if creado else 1
            messages.success(request, f"Importación: {creados} creados, {actualizados} actualizados, {omitidos} omitidos.")
            return redirect('gestion_masiva')

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, "Seleccione un archivo Excel.")
            return redirect('gestion_masiva')
        filas, errores = _leer_excel(archivo)
        if errores:
            for e in errores: messages.error(request, e)
            return redirect('gestion_masiva')

        preview = []
        for i, fila in enumerate(filas, 1):
            row = dict(fila)
            errores = []
            for campo in CAMPOS_IMPORTAR_REQUERIDOS:
                if not row.get(campo, '').strip():
                    errores.append(f"'{campo}' es requerido")
            if row.get('jornada') and row['jornada'] not in VALORES_JORNADA:
                errores.append(f"Jornada '{row['jornada']}' inválida")
            if row.get('linea') and row['linea'] not in VALORES_LINEA:
                errores.append(f"Línea '{row['linea']}' inválida")
            advertencias = []
            archivo_foto = row.get('archivo_foto', '').strip()
            if archivo_foto:
                ruta_foto = os.path.join(settings.MEDIA_ROOT, 'fotos', archivo_foto)
                if not os.path.isfile(ruta_foto):
                    advertencias.append(f"Foto '{archivo_foto}' no encontrada en la carpeta de fotos")
            doc = row.get('documento', '').strip()
            row['fila']    = i
            row['errores'] = errores
            row['advertencias'] = advertencias
            row['error']   = bool(errores)
            row['existe']  = Estudiante.objects.filter(documento=doc).exists() if doc else False
            preview.append(row)

        validos   = sum(1 for r in preview if not r['error'])
        invalidos = sum(1 for r in preview if r['error'])
        return render(request, 'estudiantes/importar_preview.html', {
            'preview': preview, 'validos': validos, 'invalidos': invalidos,
            'filas_json': json.dumps(preview),
        })
    return redirect('gestion_masiva')


@login_required
@solo_directivo
def editar_masivo(request):
    if request.method != 'POST':
        return redirect('gestion_masiva')
    jornada = request.POST.get('filtro_jornada', '')
    curso   = request.POST.get('filtro_curso', '')
    linea   = request.POST.get('filtro_linea', '')
    if not (jornada or curso or linea):
        messages.error(request, "Defina al menos un filtro.")
        return redirect('gestion_masiva')
    qs = Estudiante.objects.all()
    if jornada: qs = qs.filter(jornada=jornada)
    if curso:   qs = qs.filter(curso__iexact=curso)
    if linea:   qs = qs.filter(linea=linea)
    if not qs.exists():
        messages.warning(request, "No se encontraron estudiantes con esos filtros.")
        return redirect('gestion_masiva')
    cambios = {}
    nj = request.POST.get('nuevo_jornada', '').strip()
    nc = request.POST.get('nuevo_curso', '').strip()
    nl = request.POST.get('nuevo_linea', '').strip()
    if nj and nj in VALORES_JORNADA: cambios['jornada'] = nj
    if nc: cambios['curso'] = nc
    if nl and nl in VALORES_LINEA:   cambios['linea']   = nl
    if not cambios:
        messages.warning(request, "No se indicaron campos a cambiar.")
        return redirect('gestion_masiva')
    total = qs.count()
    qs.update(**cambios)
    messages.success(request, f"Se actualizaron {total} estudiantes.")
    return redirect('gestion_masiva')


# ── Gestión de usuarios (solo directivos) ───────────────────

@login_required
@solo_directivo
def lista_usuarios(request):
    usuarios = User.objects.exclude(is_superuser=True).select_related().prefetch_related('groups').order_by('last_name', 'first_name')
    return render(request, 'usuarios/lista.html', {
        'usuarios': usuarios,
        'es_directivo': True,
    })


@login_required
@solo_directivo
def crear_usuario(request):
    _asegurar_grupos()
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST)
        if form.is_valid():
            password_plano = form.cleaned_data['password1']
            rol = form.cleaned_data['rol']
            user = form.save(commit=False)
            user.is_staff = False
            user.save()

            # Asignar grupo
            grupo, _ = Group.objects.get_or_create(name=rol)
            user.groups.set([grupo])

            # Enviar correo de bienvenida
            resultado = _enviar_bienvenida(user, password_plano, rol)
            if resultado is True:
                messages.success(request, f"Usuario '{user.username}' creado. Se envió correo de bienvenida a {user.email}.")
            else:
                messages.warning(request, f"Usuario creado, pero no se pudo enviar el correo: {resultado}")

            return redirect('lista_usuarios')
    else:
        form = UsuarioCrearForm()

    return render(request, 'usuarios/form_usuario.html', {
        'form':  form,
        'titulo': 'Crear usuario',
        'es_nuevo': True,
    })


@login_required
@solo_directivo
def editar_usuario(request, id):
    _asegurar_grupos()
    usuario = get_object_or_404(User, id=id)
    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST, instance=usuario)
        if form.is_valid():
            rol = form.cleaned_data['rol']
            user = form.save()
            grupo, _ = Group.objects.get_or_create(name=rol)
            user.groups.set([grupo])
            messages.success(request, f"Usuario '{user.username}' actualizado.")
            return redirect('lista_usuarios')
    else:
        form = UsuarioEditarForm(instance=usuario)

    return render(request, 'usuarios/form_usuario.html', {
        'form':    form,
        'titulo':  f'Editar usuario: {usuario.get_full_name() or usuario.username}',
        'usuario': usuario,
        'es_nuevo': False,
    })


@login_required
@solo_directivo
def eliminar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)
    if usuario == request.user:
        messages.error(request, "No puede eliminar su propia cuenta.")
        return redirect('lista_usuarios')
    nombre = usuario.get_full_name() or usuario.username
    usuario.delete()
    messages.success(request, f"Usuario '{nombre}' eliminado.")
    return redirect('lista_usuarios')
